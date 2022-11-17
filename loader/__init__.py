from openpyxl.reader.excel import load_workbook
from openpyxl.workbook.workbook import Workbook
import sqlite3

from loader.model import InterviewResult
from loader.sql import(
    DEFINE_TABLE as SQL_DEFINE_TABLE,
    APPEND_ROW as SQL_APPEND_ROW
)


class StorageError(Exception):
    pass


def load_result(file_path: str, respondent_id: str) -> InterviewResult:
    book = load_workbook(file_path, read_only=True, data_only=True)
    result = InterviewResult(respondent_id)
    parse_questionary(book, result)
    parse_chronic_fatigue(book, result)
    parse_professional_burnout(book, result)
    parse_coping_strategy(book, result)
    parse_irrational_setups(book, result)
    return result


def parse_questionary(wb: Workbook, result: InterviewResult):
    sheet = wb["Анкета"]
    col = (ord('B') - ord('A')) + 1  # 1-numeration is used
    result.date = sheet.cell(row=6, column=col).value.date()
    result.gender = sheet.cell(row=7, column=col).value
    result.age = sheet.cell(row=8, column=col).value
    result.position = sheet.cell(row=9, column=col).value
    result.experience = int(sheet.cell(row=11, column=col).value)
    result.jobs_num = int(sheet.cell(row=12, column=col).value)


def parse_chronic_fatigue(wb: Workbook, result: InterviewResult):
    sheet = wb["Хроническое утомление"]
    col = (ord('D') - ord('A')) + 1  # 1-numeration is used
    result.ihru = sheet.cell(row=41, column=col).value
    result.phy_discomfort = sheet.cell(row=42, column=col).value
    result.cog_discomfort = sheet.cell(row=43, column=col).value
    result.ea_violation = sheet.cell(row=44, column=col).value
    result.motivation_dec = sheet.cell(row=45, column=col).value


def parse_professional_burnout(wb: Workbook, result: InterviewResult):
    sheet = wb["Профессиональное выгорание"]
    col = (ord('C') - ord('A')) + 1  # 1-numeration is used
    result.emotional_exhaustion = sheet.cell(row=26, column=col).value
    result.depersonalization = sheet.cell(row=27, column=col).value
    result.prof_reduction = sheet.cell(row=28, column=col).value
    result.burnout_index = sheet.cell(row=29, column=col).value


def parse_coping_strategy(wb: Workbook, result: InterviewResult):
    sheet = wb["Способы совладающего поведения"]
    col = (ord('D') - ord('A')) + 1  # 1-numeration
    result.confrontation = sheet.cell(row=56, column=col).value
    result.distancing = sheet.cell(row=57, column=col).value
    result.selfcontrol = sheet.cell(row=58, column=col).value
    result.soc_sup_search = sheet.cell(row=59, column=col).value
    result.responsibility_taking = sheet.cell(row=60, column=col).value
    result.escaping = sheet.cell(row=61, column=col).value
    result.problem_solving_planning = sheet.cell(row=62, column=col).value
    result.positive_revaluation = sheet.cell(row=63, column=col).value


def parse_irrational_setups(wb: Workbook, result: InterviewResult):
    sheet = wb["Диагностика иррациональных уста"]
    col = (ord('C') - ord('A')) + 1  # 1-numeration
    result.catastrophizing = sheet.cell(row=54, column=col).value
    result.obligation_to_self = sheet.cell(row=55, column=col).value
    result.obligation_to_others = sheet.cell(row=56, column=col).value
    result.frustration_tolerance = sheet.cell(row=57, column=col).value
    result.selfesteem = sheet.cell(row=58, column=col).value


def append_to_db(filename: str, result: InterviewResult):
    try:
        connection = create_database(filename)
    except sqlite3.DatabaseError as e:
        raise StorageError(f"DB error: {e}")
    except Exception as e:
        raise StorageError(f"Internal db error: {e}")
    cursor = connection.cursor()
    try:
        cursor.execute(SQL_APPEND_ROW, result.__dict__)
        connection.commit()
    except sqlite3.OperationalError as e:
        raise StorageError(f"SQL error: {e}")
    except sqlite3.DatabaseError as e:
        raise StorageError(f"DB error: {e}")
    except Exception as e:
        raise StorageError(f"Internal db error: {e}")


def create_database(filename: str) -> sqlite3.Connection:
    connection = sqlite3.connect(filename)
    cursor = connection.cursor()
    res = cursor.execute(SQL_DEFINE_TABLE)
    connection.commit()
    return connection
