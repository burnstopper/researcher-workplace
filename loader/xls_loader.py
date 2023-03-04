import datetime

from openpyxl.reader.excel import load_workbook
from openpyxl.workbook.workbook import Workbook

from loader.storage import ResultsTestRecord
from loader.tools import *


def load_xls_result(file_path: str, key: str, respondent_id: str) -> ResultsTestRecord:
    book = load_workbook(file_path, read_only=True, data_only=True)
    result = ResultsTestRecord()
    result.respondent_id = respondent_id
    result.key_source = key
    parse_questionary(book, result)
    parse_chronic_fatigue(book, result)
    parse_professional_burnout(book, result)
    parse_coping_strategy(book, result)
    parse_irrational_setups(book, result)
    return result


def parse_questionary(wb: Workbook, result: ResultsTestRecord):
    sheet = wb["Анкета"]
    col = (ord('B') - ord('A')) + 1  # 1-numeration is used
    filling_time: datetime.datetime = sheet.cell(row=6, column=col).value
    result.gender = recode_gender(sheet.cell(row=7, column=col).value)
    result.age = sheet.cell(row=8, column=col).value
    result.age_c = encode_age_cat(result.age)
    result.position = sheet.cell(row=9, column=col).value
    result.experience = int(sheet.cell(row=11, column=col).value)
    result.jobs_num = int(sheet.cell(row=12, column=col).value)
    result.date_time = filling_time - datetime.timedelta(filling_time.microsecond)


def parse_chronic_fatigue(wb: Workbook, result: ResultsTestRecord):
    sheet = wb["Хроническое утомление"]
    col = (ord('D') - ord('A')) + 1  # 1-numeration is used
    col_p = col + 1
    result.distress = sheet.cell(row=41, column=col).value
    result.distress_c = distress_to_cat(result.distress)
    result.distress_p = float(sheet.cell(row=41, column=col_p).value)
    result.physical_discomfort = sheet.cell(row=42, column=col).value
    result.physical_discomfort_p = float(sheet.cell(row=42, column=col_p).value)
    result.cognitive_discomfort = sheet.cell(row=43, column=col).value
    result.cognitive_discomfort_p = float(sheet.cell(row=43, column=col_p).value)
    result.ea_violation = sheet.cell(row=44, column=col).value
    result.ea_violation_p = float(sheet.cell(row=44, column=col_p).value)
    result.motivation_decrease = sheet.cell(row=45, column=col).value
    result.motivation_decrease_p = float(sheet.cell(row=45, column=col_p).value)


def parse_professional_burnout(wb: Workbook, result: ResultsTestRecord):
    sheet = wb["Профессиональное выгорание"]
    col = (ord('C') - ord('A')) + 1  # 1-numeration is used
    col_p = col + 2
    result.emotional_exhaustion = sheet.cell(row=26, column=col).value
    result.emotional_exhaustion_p = float(sheet.cell(row=26, column=col_p).value)
    result.emotional_exhaustion_c = emotional_exhaustion_to_cat(sheet.cell(row=26, column=col).value)
    result.depersonalization = sheet.cell(row=27, column=col).value
    result.depersonalization_p = float(sheet.cell(row=27, column=col_p).value)
    result.depersonalization_c = depersonalization_to_cat(sheet.cell(row=27, column=col).value)
    result.reduction_of_professionalism = sheet.cell(row=28, column=col).value
    result.reduction_of_professionalism_c = reduction_of_professionalism_to_cat(sheet.cell(row=28, column=col).value)
    result.reduction_of_professionalism_p = float(sheet.cell(row=28, column=col_p).value)
    result.burnout_p = sheet.cell(row=29, column=col).value
    result.burnout_index = \
        result.emotional_exhaustion \
        + result.depersonalization \
        + MAX_REDUCTION_OF_PROFESSIONALISM \
        - result.reduction_of_professionalism
    result.burnout_index_p = result.burnout_index / (MAX_EMOTIONAL_EXHAUSTION + MAX_DEPERSONALIZATION + MAX_REDUCTION_OF_PROFESSIONALISM)


def parse_coping_strategy(wb: Workbook, result: ResultsTestRecord):
    sheet = wb["Способы совладающего поведения"]
    col = (ord('D') - ord('A')) + 1  # 1-numeration
    result.confrontation = sheet.cell(row=56, column=col).value
    result.confrontation_c = lazarus_to_cat(result.confrontation)
    result.distancing = sheet.cell(row=57, column=col).value
    result.distancing_c = lazarus_to_cat(result.distancing)
    result.selfcontrol = sheet.cell(row=58, column=col).value
    result.selfcontrol_c = lazarus_to_cat(result.selfcontrol)
    result.seeking_social_support = sheet.cell(row=59, column=col).value
    result.seeking_social_support_c = lazarus_to_cat(result.seeking_social_support)
    result.taking_responsibility = sheet.cell(row=60, column=col).value
    result.taking_responsibility_c = lazarus_to_cat(sheet.cell(row=60, column=col).value)
    result.escaping = sheet.cell(row=61, column=col).value
    result.escaping_c = lazarus_to_cat(result.escaping)
    result.problem_solving_planning = sheet.cell(row=62, column=col).value
    result.problem_solving_planning_c = lazarus_to_cat(result.problem_solving_planning)
    result.positive_revaluation = sheet.cell(row=63, column=col).value
    result.positive_revaluation_c = lazarus_to_cat(result.positive_revaluation)


def parse_irrational_setups(wb: Workbook, result: ResultsTestRecord):
    sheet = wb["Диагностика иррациональных уста"]
    col = (ord('C') - ord('A')) + 1  # 1-numeration
    result.catastrophization = sheet.cell(row=54, column=col).value
    result.catastrophization_c = spb_to_cat(sheet.cell(row=54, column=col).value)
    result.due_to_self = sheet.cell(row=55, column=col).value
    result.due_to_self_c = spb_to_cat(sheet.cell(row=55, column=col).value)
    result.due_to_others = sheet.cell(row=56, column=col).value
    result.due_to_others_c = spb_to_cat(sheet.cell(row=56, column=col).value)
    result.low_frustration_tolerance = sheet.cell(row=57, column=col).value
    result.low_frustration_tolerance_c = spb_to_cat(sheet.cell(row=57, column=col).value)
    result.selfestimation = sheet.cell(row=58, column=col).value
    result.selfestimation_c = spb_to_cat(sheet.cell(row=58, column=col).value)
